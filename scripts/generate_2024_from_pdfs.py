#!/usr/bin/env python3
"""
generate_2024_from_pdfs.py

Script to extract tables from monthly PDFs in data_2024/ and produce .xlsx files in output_2024/
using the 2023 monthly .xlsx files in examples_2023_split/ as templates for layout/formatting.

Usage:
  - Install dependencies: pip install pdfplumber pandas openpyxl xlrd
  - From repo root run: python3 scripts/generate_2024_from_pdfs.py

Notes/limitations:
  - This script copies the corresponding 2023 template (if present) to output_2024 and then
    writes extracted table rows starting at a configurable start_row (default 4) in the first sheet.
    Copying the template preserves formatting; the data is written into the cells preserving styles.
  - Extraction heuristics may need adjustments per PDF layout. Test on a single month first.
"""

import os
import glob
import shutil
from pathlib import Path
import pdfplumber
import pandas as pd
from openpyxl import load_workbook

# Configuration
REPO_ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = REPO_ROOT / 'data_2024'
TEMPLATE_DIR = REPO_ROOT / 'examples_2023_split'
OUTPUT_DIR = REPO_ROOT / 'output_2024'
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# Where to start writing extracted table into the template workbook (1-indexed)
WRITE_START_ROW = 4
WRITE_START_COL = 1  # column A

# Mapping: filenames in data_2024 are like GENNAIO_2024.pdf, etc.
# We'll derive month name from the filename before the first underscore.

def extract_tables_from_pdf(pdf_path):
    """Extract tables from all pages of a pdf and return a single DataFrame.
    Uses pdfplumber's extract_table per page and concatenates.
    """
    tables = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            try:
                page_tables = page.extract_tables()
            except Exception:
                page_tables = None
            if not page_tables:
                continue
            for t in page_tables:
                # convert table (list of rows) to DataFrame
                if not t:
                    continue
                df = pd.DataFrame(t)
                # if first row looks like header (no None values), set as header
                if df.shape[0] >= 2 and df.iloc[0].notnull().all():
                    df.columns = df.iloc[0]
                    df = df.drop(df.index[0]).reset_index(drop=True)
                tables.append(df)
    if not tables:
        return pd.DataFrame()
    # normalize columns by taking union and concatenating
    try:
        result = pd.concat(tables, ignore_index=True, sort=False)
    except ValueError:
        # fallback: concat by rows with reset columns
        result = pd.concat([t.astype(str) for t in tables], ignore_index=True, sort=False)
    # clean up empty columns/rows
    result = result.dropna(how='all')
    result = result.loc[:, result.columns.notnull()]
    return result


def write_df_into_workbook(workbook_path, df, start_row=WRITE_START_ROW, start_col=WRITE_START_COL):
    """Open workbook, write df into first sheet starting at start_row, start_col, then save workbook.
    This will preserve existing formatting in the template.
    """
    wb = load_workbook(workbook_path)
    ws = wb[wb.sheetnames[0]]

    # Clear previous data in the area where we'll write (simple heuristic: clear rows from start_row to start_row+len(df)+10)
    max_rows_to_clear = max(50, len(df) + 10)
    for r in range(start_row, start_row + max_rows_to_clear):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            if cell.value is not None:
                cell.value = None

    # Write DataFrame values preserving no styles for new values (styles in template row will remain for other cells).
    for i, row in enumerate(df.itertuples(index=False), start=0):
        for j, value in enumerate(row, start=0):
            ws.cell(row=start_row + i, column=start_col + j, value=value)

    wb.save(workbook_path)


def process_month_pdf(pdf_path):
    pdf_name = Path(pdf_path).name
    month_key = pdf_name.split('_')[0]  # e.g., 'MARZO' from MARZO_2024.pdf
    template_name = f"{month_key} 2023.xlsx"
    template_path = TEMPLATE_DIR / template_name
    if not template_path.exists():
        # fallback to generic template
        template_path = TEMPLATE_DIR / '2023.xlsx'
        if not template_path.exists():
            raise FileNotFoundError(f"No template found for {month_key} in {TEMPLATE_DIR} and no generic 2023.xlsx")

    out_filename = f"{month_key} 2024.xlsx"
    out_path = OUTPUT_DIR / out_filename

    # Copy template to output
    shutil.copyfile(template_path, out_path)
    print(f"Copied template {template_path.name} -> {out_path}")

    # Extract tables
    df = extract_tables_from_pdf(pdf_path)
    if df.empty:
        print(f"No tables found in {pdf_name}; leaving template copy as-is.")
        return out_path

    # Attempt to drop fully-empty columns
    df = df.dropna(axis=1, how='all')
    # Reset columns to simple strings if they are tuples or other types
    df.columns = [str(c) for c in df.columns]

    # Write into workbook
    write_df_into_workbook(str(out_path), df)
    print(f"Wrote extracted table to {out_path}")
    return out_path


def main():
    pdf_files = sorted(glob.glob(str(DATA_DIR / '*_2024.pdf')))
    if not pdf_files:
        print(f"No monthly PDFs found in {DATA_DIR}")
        return

    created = []
    for pdf in pdf_files:
        try:
            out = process_month_pdf(pdf)
            created.append(out)
        except Exception as e:
            print(f"Error processing {pdf}: {e}")

    print(f"Done. Created {len(created)} files in {OUTPUT_DIR}")


if __name__ == '__main__':
    main()
