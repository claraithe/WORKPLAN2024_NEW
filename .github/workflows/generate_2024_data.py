import os
import datetime
import pandas as pd
import pdfplumber
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

DATA_DIR = "data_2024"
CSV_FILE = "report (29).csv"
PDF_MONTHS = ["GENNAIO_2024.pdf","FEBBRAIO_2024.pdf","MARZO_2024.pdf","APRILE_2024.pdf",
              "MAGGIO_2024.pdf","GIUGNO_2024.pdf","LUGLIO_2024.pdf","AGOSTO_2024.pdf",
              "SETTEMBRE_2024.pdf","OTTOBRE_2024.pdf","NOVEMBRE_2024.pdf","DICEMBRE_2024.pdf"]
OUTPUT_DIR = "output_2024"
TRACE_MD_FILE = "TRACE_2024.md"

def date_range(year):
    start = datetime.date(year, 1, 1)
    end = datetime.date(year, 12, 31)
    for n in range((end - start).days + 1):
        yield start + datetime.timedelta(days=n)

def parse_csv(csv_path):
    df = pd.read_csv(csv_path)
    date_col = next((c for c in df.columns if "date" in c.lower()), df.columns[0])
    df['Date'] = pd.to_datetime(df[date_col], dayfirst=True, errors='coerce').dt.date
    flights = {row['Date']: row for _, row in df.iterrows() if pd.notnull(row['Date'])}
    return flights

def parse_pdfs(pdf_dir):
    code_by_date = {}
    for fname in PDF_MONTHS:
        fpath = os.path.join(pdf_dir, fname)
        if not os.path.exists(fpath): continue
        with pdfplumber.open(fpath) as pdf:
            for page in pdf.pages:
                table = page.extract_table()
                if not table: continue
                for row in table:
                    if not row or not row[0]: continue
                    try:
                        dt = pd.to_datetime(row[0], dayfirst=True, errors='coerce').date()
                        code = str(row[1]).strip() if len(row) > 1 and row[1] else ""
                        if dt and code:
                            code_by_date[dt] = code
                    except Exception:
                        continue
    return code_by_date

def build_consolidated(flights, codes):
    consolidated = []
    for day in date_range(2024):
        entry = {'Date': day.strftime("%Y-%m-%d")}
        # 1. CSV - Volo
        if day in flights:
            entry['Type'] = "Volo"
            for k, v in flights[day].items():
                entry[k] = v
        # 2. PDF Duty Codes
        elif day in codes:
            entry['Type'] = codes[day]
        # 3. OFF (auto)
        else:
            entry['Type'] = "OFF (auto)"
        consolidated.append(entry)
    return pd.DataFrame(consolidated)

def overwrite_excel_files(consolidated_df, out_dir):
    excel_files = [f for f in os.listdir(out_dir) if f.lower().endswith('.xlsx')]
    for fname in excel_files:
        fpath = os.path.join(out_dir, fname)
        wb = load_workbook(fpath)
        ws = wb.active
        # Overwrite content (assume header in row 1, start writing from row 2)
        header = [c.value for c in ws[1]]
        for i, row in enumerate(dataframe_to_rows(consolidated_df, index=False, header=False), start=2):
            for j, value in enumerate(row):
                if j < len(header):
                    ws.cell(row=i,column=j+1).value = value
        wb.save(fpath)

def generate_trace_md(consolidated_df, out_dir, filename):
    months = {m: [] for m in range(1, 13)}
    for _, row in consolidated_df.iterrows():
        dt = datetime.datetime.strptime(row['Date'], "%Y-%m-%d").date()
        months[dt.month].append(row)
    with open(os.path.join(out_dir, filename), "w", encoding="utf-8") as f:
        for m in range(1, 13):
            f.write(f"# {datetime.date(2024, m, 1).strftime('%B').upper()} 2024\n")
            vo, pdf, auto_off = 0, 0, 0
            for r in months[m]:
                t = r['Type']
                if t == "Volo": vo += 1
                elif t == "OFF (auto)": auto_off += 1
                else: pdf += 1
                f.write(f"- {r['Date']}: {t}\n")
            f.write(f"\nSummary: Volo={vo}, PDF codes={pdf}, OFF (auto)={auto_off}\n\n")

def main():
    print("Parsing flight CSV...")
    flights = parse_csv(os.path.join(DATA_DIR, CSV_FILE))
    print("Parsing monthly PDF duty logs...")
    codes = parse_pdfs(DATA_DIR)
    print("Building consolidated 2024 table...")
    consolidated_df = build_consolidated(flights, codes)
    print("Filling Excel output files...")
    overwrite_excel_files(consolidated_df, OUTPUT_DIR)
    print("Generating monthly summary markdown...")
    generate_trace_md(consolidated_df, OUTPUT_DIR, TRACE_MD_FILE)
    print("Done! All outputs regenerated with real 2024 data.")

if __name__ == "__main__":
    main()
